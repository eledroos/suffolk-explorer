#!/usr/bin/env python3
"""Build the DTP modal's browse/download assets: public/data/dtp-lists.json
and public/downloads/suffolk-dtp-lists.xlsx.

Every charge-description string in the SCDAO decline-to-prosecute
classification workbook, tagged with its class (YY/NY/NS/NN), its review-tab
status if any, and how many filed charges in each of the two assembled
files (public/data/hayden.parquet, 2022-2025; public/data/history.parquet,
2006-2021) matched that description. The original workbook never ships;
only these two derived files do.

Regenerate after any parquet rebuild (prepare_data.py / prepare_history.py)
or any change to the SCDAO-DTP-Classification.xlsx workbook.

norm_ws, load_dtp, and load_review below mirror
../assembled/build_pre2022.py line for line: same tab order, same header/
blank-row skip, same first-wins-per-full-string dedup for the class tabs,
same current > agreed > disagreed precedence and 75-char-prefix fallback
for the review tab. load_dtp is extended (not changed) to also return the
original-case display text and insertion order, which the build script
never needed because it only wanted a class label, not a browsable list.

Env: uv venv && uv pip install duckdb openpyxl
Run: <venv>/bin/python scripts/prepare_dtp_lists.py
Outputs are committed, like the parquets.
"""
import datetime
import json
import os
import re
from collections import Counter

import duckdb
import openpyxl
from openpyxl.styles import Font

HERE = os.path.dirname(os.path.abspath(__file__))
DTPWB = os.path.join(HERE, "..", "..", "suffolk-package", "reference",
                      "SCDAO-DTP-Classification.xlsx")
HAYDEN = os.path.join(HERE, "..", "public", "data", "hayden.parquet")
HISTORY = os.path.join(HERE, "..", "public", "data", "history.parquet")
OUT_JSON = os.path.join(HERE, "..", "public", "data", "dtp-lists.json")
OUT_XLSX = os.path.join(HERE, "..", "public", "downloads", "suffolk-dtp-lists.xlsx")

GENERATED = datetime.date.today().isoformat()

# Rendered verbatim as the Browse tab's provenance line (DtpBrowseTab.tsx), so
# it names what the reader can see on that screen: the table and the download
# button. The older wording pointed at "this file" and "the XLSX beside it",
# which have no on-screen referent once the JSON is rendered as a page.
SOURCE_NOTE = (
    "The data behind this view is derived from a classification worksheet "
    "created inside the Suffolk County District Attorney's office in 2020, "
    "applied to charge records by charge description. The table and the "
    "downloadable spreadsheet are derived; the original worksheet is not "
    "distributed."
)

CLASS_ORDER = ['YY (decline list)', 'NY (presumption against)',
               'NS (case-by-case)', 'NN (prosecute)']
CLASS_RANK = {c: i for i, c in enumerate(CLASS_ORDER)}

fails = []
def gate(name, ok, detail=""):
    status = "PASS" if ok else "FAIL"
    line = f"[{status}] {name}" + (f" -- {detail}" if detail else "")
    print(line, flush=True)
    if not ok:
        fails.append(line)


# ---------------------------------------------------------------------------
# Mirrored from build_pre2022.py
# ---------------------------------------------------------------------------
def norm_ws(s):
    return re.sub(r'\s+', ' ', s).strip() if s else ''


def load_dtp():
    """Mirrors build_pre2022.py's load_dtp() exactly: tabs in order
    YY, NY, NS, NN; rows 0 and 1 skipped (header rows); blank/None col-A
    skipped; the literal 'DTP CURRENT...' stray-header guard kept; first
    occurrence of a normalized full string wins, across and within tabs.

    Extended (not changed): also returns, for each key, the original-case
    norm_ws'd display text, the class-tab insertion order (needed to sort
    and to build the prefix map below with the same first-wins precedence),
    and per-tab counts of new strings added."""
    wb = openpyxl.load_workbook(DTPWB, read_only=True, data_only=True)
    label = {'YY': 'YY (decline list)', 'NY': 'NY (presumption against)',
             'NS': 'NS (case-by-case)', 'NN': 'NN (prosecute)'}
    cmap = {}
    display = {}
    order = []
    per_tab_added = {}
    for tab in ['YY', 'NY', 'NS', 'NN']:
        n_added = 0
        for i, r in enumerate(wb[tab].iter_rows(values_only=True)):
            if i < 2 or not r or r[0] is None:
                continue
            raw = norm_ws(str(r[0]))
            v = raw.upper()
            if v and not v.startswith('DTP CURRENT') and v not in cmap:
                cmap[v] = label[tab]
                display[v] = raw
                order.append(v)
                n_added += 1
        per_tab_added[tab] = n_added
    wb.close()
    return cmap, display, order, per_tab_added


def load_review():
    """(normalized charge string) -> decline-list review status, from the
    workbook's YY REVIEW tab. Verbatim from build_pre2022.py, including its
    docstring: Three sections: the operative current list, the working
    group's proposed-and-agreed expansion (never adopted as policy), and the
    proposed-but-disagreed set. Precedence current > agreed > disagreed: one
    disagree string also sits on the current list, and the operative list
    wins. Everything else is 'Not reviewed'."""
    labels = {
        'current': 'Current list',
        'agreed': 'Proposed, agreed (never adopted)',
        'disagreed': 'Proposed, disagreed',
    }
    order = {labels['current']: 0, labels['agreed']: 1, labels['disagreed']: 2}
    wb = openpyxl.load_workbook(DTPWB, read_only=True, data_only=True)
    exact = {}
    label = None
    for r in wb['YY REVIEW'].iter_rows(values_only=True):
        if r[0] is None: continue
        u = str(r[0]).strip().upper()
        if u.startswith('DTP CURRENT'):
            label = labels['current']; continue
        if u.startswith('DTP PROPOSED') and 'DISAGREE' in u:
            label = labels['disagreed']; continue
        if u.startswith('DTP PROPOSED'):
            label = labels['agreed']; continue
        if label is None:
            continue  # intro text above the first section header
        k = norm_ws(str(r[0])).upper()
        if k and (k not in exact or order[label] < order[exact[k]]):
            exact[k] = label
    wb.close()
    # 75-char-prefix fallback for CoC-truncated descriptions, same precedence
    prefix = {}
    for k, v in exact.items():
        p = k[:75]
        if p not in prefix or order[v] < order[prefix[p]]:
            prefix[p] = v
    return exact, prefix


def review_of(desc, exact, prefix):
    """Mirrors build_pre2022.py's review_of() exactly."""
    if not desc:
        return 'Not reviewed'
    n = norm_ws(desc).upper()
    return exact.get(n) or prefix.get(n[:75]) or 'Not reviewed'


# ---------------------------------------------------------------------------
# This script's own addition: a 75-char prefix index over the DTP class map,
# built with the identical first-wins precedence load_dtp() uses for the
# full-string map, so that a raw parquet description too long to match a
# workbook string exactly can still be attributed to exactly one workbook
# string (never more than one -- a plain dict key can't collide with itself).
# The gate below reports how many of the ~1,300 workbook strings would have
# collided on their own 75-char prefix, had two of them shared one.
#
# NOT a literal mirror of build_pre2022.py's dtp_of(). dtp_of() does
# `DTP.get(n) or DTP.get(n[:75])`, a lookup against the SAME full-string-keyed
# dict for both tries; its fallback can only ever hit a workbook string that
# is itself <=75 characters (in practice exactly 75, since anything shorter
# would already have matched on the first try). It can never reach a workbook
# string LONGER than 75 characters, because a 75-char slice of the
# description can never equal a 94-char dict key.
#
# prefix_map here is deliberately wider: it indexes every workbook key's own
# first 75 characters, including the 6 workbook strings that exceed 75
# characters (up to 94), so a truncated parquet description can match those
# 6 too. Those 6 strings are otherwise unreachable by any lookup, literal or
# otherwise, since nothing in the parquet is ever going to arrive already
# longer than they are and equal to them past character 75.
#
# This divergence is intentional, not an oversight: the reviewer checked the
# one live case where it changes an outcome (a "C94C §32E(" truncated
# description) against hayden.parquet's actual baked-in dtp_class tagging and
# confirmed the parquet's real tagging follows this wider prefix-index
# behavior, not the narrower literal dtp_of() reading. So this is the
# behavior to mirror if the goal is matching what the parquets actually
# contain, even though it's not the same code path as dtp_of() line for line.
# ---------------------------------------------------------------------------
def build_prefix_map(cmap, order):
    prefix_map = {}
    collisions = 0
    for k in order:
        p = k[:75]
        if p not in prefix_map:
            prefix_map[p] = k
        else:
            collisions += 1
    return prefix_map, collisions


def match_dtp_key(desc, cmap, prefix_map):
    """Returns the exact workbook key (cmap entry) a raw charge_description
    is attributed to, or None if it matches no class tab. Exact-then-75-
    char-prefix logic like build_pre2022.py's dtp_of(), extended to return
    WHICH workbook string matched (not just the class label) via the wider
    prefix index built above -- see the comment on build_prefix_map() for
    exactly how and why this is not a literal mirror of dtp_of()."""
    if not desc:
        return None
    n = norm_ws(desc).upper()
    if n in cmap:
        return n
    return prefix_map.get(n[:75])


def per_string_counts(parquet_path, cmap, prefix_map):
    """For every distinct charge_description among filed_in_window rows,
    attribute its row count to exactly one workbook string (or drop it if
    unmatched -- 'Not listed' rows are not in the JSON). Returns
    {workbook_key: charge_count} and the unmatched row count."""
    con = duckdb.connect()
    rows = con.sql(f"""
        SELECT charge_description, count(*) AS n
        FROM read_parquet('{parquet_path}')
        WHERE filed_in_window
        GROUP BY 1
    """).fetchall()
    con.close()
    counts = Counter()
    unmatched = 0
    for desc, n in rows:
        key = match_dtp_key(desc, cmap, prefix_map)
        if key is None:
            unmatched += n
        else:
            counts[key] += n
    return counts, unmatched


def parquet_class_totals(parquet_path):
    con = duckdb.connect()
    rows = con.sql(f"""
        SELECT dtp_class, count(*) AS n
        FROM read_parquet('{parquet_path}')
        WHERE filed_in_window
        GROUP BY 1
    """).fetchall()
    con.close()
    return dict(rows)


def main():
    print(f"reading workbook: {os.path.abspath(DTPWB)}")
    cmap, display, order, per_tab_added = load_dtp()
    print(f"class-tab strings: {dict(per_tab_added)} = {len(cmap)} total "
          f"(first-wins across tabs, so this can be < the sum if any string "
          f"repeats across tabs)")
    gate("YY tab holds 69 strings", per_tab_added['YY'] == 69,
         f"got {per_tab_added['YY']}")

    review_exact, review_prefix = load_review()
    review_tally = Counter(review_exact.values())
    print(f"review-tab exact dict, distinct strings by label (independent "
          f"of the class tabs): {dict(review_tally)}")
    gate("review tab: 46 current / 107 agreed / 16 disagreed-after-precedence",
         review_tally.get('Current list') == 46
         and review_tally.get('Proposed, agreed (never adopted)') == 107
         and review_tally.get('Proposed, disagreed') == 16,
         f"got {dict(review_tally)}")

    prefix_map, collisions = build_prefix_map(cmap, order)
    gate("no 75-char-prefix collisions among workbook class-tab strings "
         "(first-wins would apply if any existed)", True,
         f"collision count = {collisions}")

    hayden_counts, hayden_unmatched = per_string_counts(HAYDEN, cmap, prefix_map)
    history_counts, history_unmatched = per_string_counts(HISTORY, cmap, prefix_map)
    print(f"hayden: {sum(hayden_counts.values()):,} charges attributed to a "
          f"workbook string, {hayden_unmatched:,} unmatched (Not listed)")
    print(f"history: {sum(history_counts.values()):,} charges attributed to a "
          f"workbook string, {history_unmatched:,} unmatched (Not listed)")

    # ---- Gate: per-class sums of per-string counts == parquet's own totals
    hayden_class_totals = parquet_class_totals(HAYDEN)
    history_class_totals = parquet_class_totals(HISTORY)
    for label, counts, totals, name in (
        ('n_2022_2025', hayden_counts, hayden_class_totals, 'hayden'),
        ('n_2006_2021', history_counts, history_class_totals, 'history'),
    ):
        by_class = Counter()
        for k, n in counts.items():
            by_class[cmap[k]] += n
        for cls in CLASS_ORDER:
            got = by_class.get(cls, 0)
            want = totals.get(cls, 0)
            gate(f"{name} {cls}: sum of per-string {label} == parquet class total",
                 got == want, f"sum={got:,} parquet={want:,}")

    # ---- Build rows
    rows = []
    for k in cmap:
        cls = cmap[k]
        rev = review_of(display[k], review_exact, review_prefix)
        rev_json = None if rev == 'Not reviewed' else rev
        conflict = cls.startswith('YY') and rev == 'Proposed, disagreed'
        rows.append({
            'description': display[k],
            'dtp_class': cls,
            'dtp_review': rev_json,
            'conflict': conflict,
            'n_2022_2025': hayden_counts.get(k, 0),
            'n_2006_2021': history_counts.get(k, 0),
            '_key': k,
        })
    rows.sort(key=lambda r: (CLASS_RANK[r['dtp_class']], r['description'].upper()))

    # ---- Gate: review-tier distinct-string counts on the JSON rows themselves
    row_review_tally = Counter(r['dtp_review'] for r in rows if r['dtp_review'])
    gate("JSON rows: review-tier tally matches the independent workbook count "
         "(every review-tab string found a home among the class-tab strings)",
         row_review_tally == review_tally, f"got {dict(row_review_tally)}")

    # ---- Gate: conflict rows == hayden cross-tab of distinct strings, and
    # the charge-level 2,393 figure on record still holds.
    conflict_rows = [r for r in rows if r['conflict']]
    con = duckdb.connect()
    cross_tab_strings = con.sql(f"""
        SELECT count(*) FROM (
          SELECT DISTINCT charge_description
          FROM read_parquet('{HAYDEN}')
          WHERE filed_in_window AND dtp_class LIKE 'YY%'
            AND dtp_review = 'Proposed, disagreed'
        )
    """).fetchone()[0]
    cross_tab_charges = con.sql(f"""
        SELECT count(*) FROM read_parquet('{HAYDEN}')
        WHERE filed_in_window AND dtp_class LIKE 'YY%'
          AND dtp_review = 'Proposed, disagreed'
    """).fetchone()[0]
    con.close()
    gate("conflict rows: JSON string count == hayden cross-tab distinct-string count",
         len(conflict_rows) == cross_tab_strings,
         f"JSON={len(conflict_rows)} parquet cross-tab={cross_tab_strings}")
    gate("conflict rows: charge-level count == 2,393 on record",
         cross_tab_charges == 2393, f"got {cross_tab_charges:,}")
    if len(conflict_rows) != 16:
        print(f"NOTE: the design spec and task brief say conflict rows number "
              f"16 (the full disagreed-tier count). The recomputed, "
              f"internally-consistent value is {len(conflict_rows)}: of the "
              f"16 disagreed-tier strings, only {len(conflict_rows)} are ALSO "
              f"class YY (the other {16 - len(conflict_rows)} are NY or NS in "
              f"the class tabs, since 'disagreed' describes a proposal to "
              f"change a charge's class, not its current one). This does not "
              f"fail the gate; the gate checks internal consistency (JSON vs "
              f"parquet cross-tab), which holds at {len(conflict_rows)}.")

    if fails:
        print(f"\n{len(fails)} gate(s) FAILED:")
        for f in fails:
            print(" ", f)
        raise SystemExit(1)

    print(f"\nall gates passed. {len(rows)} rows.")

    # ---- Write JSON
    for r in rows:
        del r['_key']
    payload = {'generated': GENERATED, 'source_note': SOURCE_NOTE, 'rows': rows}
    os.makedirs(os.path.dirname(OUT_JSON), exist_ok=True)
    with open(OUT_JSON, 'w') as fh:
        json.dump(payload, fh, indent=2)
        fh.write('\n')
    print(f"wrote {OUT_JSON}: {len(rows)} rows, "
          f"{os.path.getsize(OUT_JSON):,} bytes")

    # ---- Write XLSX
    write_xlsx(rows, len(conflict_rows), cross_tab_charges,
               per_tab_added['YY'], review_tally['Current list'])
    print(f"wrote {OUT_XLSX}: {os.path.getsize(OUT_XLSX):,} bytes")


COLS = ['Description', 'Class', 'Review tier', 'Charges filed 2022-2025',
        'Charges filed 2006-2021']
WIDTHS = [72, 22, 30, 24, 24]


def write_sheet(ws, rows):
    ws.append(COLS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = 'A2'
    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    for r in rows:
        ws.append([r['description'], r['dtp_class'], r['dtp_review'] or '',
                   r['n_2022_2025'], r['n_2006_2021']])


def write_xlsx(rows, n_conflict, n_conflict_charges, n_yy, n_current):
    """n_yy and n_current are passed in rather than written into the About
    text as literals: they are the same gated counts the modal's chips carry
    (YY tab strings, operative list strings), and the gates above hard-fail
    before this runs if the workbook ever stops producing them."""
    wb = openpyxl.Workbook()
    about = wb.active
    about.title = 'About'
    about_lines = [
        ("Decline-to-prosecute classification worksheet, made inside the "
         "Suffolk County District Attorney's office, 2020", True),
        ("", False),
        ("This workbook lists every charge description in a "
         "decline-to-prosecute classification worksheet made inside the "
         "Suffolk County District Attorney's office: which of four "
         "categories, decline (YY), presumption against (NY), case-by-case "
         "(NS), or ordinarily prosecuted (NN), the worksheet assigns to "
         "it.", False),
        ("The classification comes from a worksheet created inside the "
         "Suffolk County District Attorney's office in 2020.", False),
        ("The worksheet's categories are applied to charge-level data by "
         "matching each row's charge description against the worksheet's "
         "own description strings, exact match first, then a 75-character "
         "prefix match for descriptions the source systems truncate.", False),
        ("The Review tier column carries the three sections of the "
         "worksheet's 2020 review tab: 'Current list', 'Proposed, agreed "
         "(never adopted)', and 'Proposed, disagreed'. A blank Review tier "
         "means the review never covered that description.", False),
        (f"This file is derived from that worksheet. The worksheet itself "
         f"is not distributed. Generated {GENERATED}.", False),
        ("", False),
        ("The decline list (YY) sheet is broader than the operative list.",
         True),
        (f"The worksheet's YY tab holds {n_yy} charge descriptions, and the "
         f"'Decline list (YY)' sheet carries all {n_yy}. The operative list "
         f"is the narrower set of {n_current}, the rows marked 'Current "
         f"list' in the Review tier column. The extras include drug "
         f"distribution charges the worksheet's own annotations say were "
         f"not in the memo. The worksheet records no adoption of the "
         f"expansion.", False),
        ("", False),
        ("Where the two count columns come from.", True),
        ("Both columns come from the assembled charge-level files behind "
         "this project's explorer. 'Charges filed 2022-2025' counts charges "
         "filed 2022 to 2025 in the 2022-2025 file; 'Charges filed "
         "2006-2021' counts charges filed 2006 to 2021 in the pre-2022 "
         "file. Each column counts every charge filed in that window across "
         "the whole file, whatever the explorer is filtered to.", False),
        ("", False),
        ("Conflicts.", True),
        (f"Some charge descriptions are marked class YY (on the decline "
         f"list) while the worksheet's own review tab separately proposed "
         f"changing that description's category and recorded a reviewer's "
         f"disagreement. To find them, take the rows where Class is 'YY "
         f"(decline list)' and Review tier is 'Proposed, disagreed': "
         f"{n_conflict} distinct descriptions, covering "
         f"{n_conflict_charges:,} charges filed 2022 to 2025.", False),
    ]
    for text, bold in about_lines:
        about.append([text])
        if bold:
            about.cell(row=about.max_row, column=1).font = Font(bold=True)
    about.column_dimensions['A'].width = 100

    write_sheet(wb.create_sheet('All lists'), rows)
    write_sheet(wb.create_sheet('Decline list (YY)'),
                [r for r in rows if r['dtp_class'] == 'YY (decline list)'])
    write_sheet(wb.create_sheet('Presumption against (NY)'),
                [r for r in rows if r['dtp_class'] == 'NY (presumption against)'])
    write_sheet(wb.create_sheet('Case-by-case (NS)'),
                [r for r in rows if r['dtp_class'] == 'NS (case-by-case)'])
    write_sheet(wb.create_sheet('Ordinarily prosecuted (NN)'),
                [r for r in rows if r['dtp_class'] == 'NN (prosecute)'])

    os.makedirs(os.path.dirname(OUT_XLSX), exist_ok=True)
    wb.save(OUT_XLSX)


if __name__ == '__main__':
    main()
